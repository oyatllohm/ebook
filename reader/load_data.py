import openpyxl
from django.conf import settings


# Define variable to load the wookbook
wookbook = openpyxl.load_workbook(f"{settings.BASE_DIR}/reader/student_list.xlsx")
worksheet = wookbook.active


def load_data():
    students = []
    for col in worksheet.iter_rows(1, worksheet.max_row):
        fullname = col[1].value 
        data = {   
        'passport':col[2].value ,
        'pinfl' :col[3].value ,
        'course': col[11].value ,
        'faculty':col[12].value ,
        'group':col[13].value ,
        }

        split_data = fullname.split()
        data['name'] = split_data[1]
        data['surname'] = split_data[0]
        data['middlename'] = " ".join(split_data[2:])
        students.append(data)
    return students    
        

